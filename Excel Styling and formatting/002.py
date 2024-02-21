import openpyxl
wb = openpyxl.load_workbook("001.xlsx")

sh3=wb['Sheet3']
row=sh3.max_row
column=sh3.max_column

#print(row)
#print(column)

for i in range(1,row+1):
    for j in range(1,column+1):
        print(sh3.cell(i,j).value)

sh3.cell(row=6,column=1,value='simggoo')
wb.save("001.xlsx")
#to creat a new saved file use wb.save("002.xlsx")