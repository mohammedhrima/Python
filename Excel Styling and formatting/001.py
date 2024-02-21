import openpyxl
wb = openpyxl.load_workbook("001.xlsx")

sheets =wb.sheetnames
print("------------------------------------")
print(sheets)
print("------------------------------------")
print(wb.active.title)
print("------------------------------------")
sh1=wb['Sheet1']

#see value in tables in xsls file option 1
Value_1 = sh1['B1'].value
print(Value_1)
print("------------------------------------")
Value_2 = wb["Sheet2"]["A2"].value
print(Value_2)
print("------------------------------------")

#see value in tables in xsls file option 2
sh3 =wb['Sheet3']
Value_3=sh3.cell(3,2).value #row1 , column2
print(Value_3)
print("------------------------------------")

#see value in tables in xsls file option 3
print(sh3.cell(row=2,column=2).value)
print("------------------------------------")
print(wb.get_sheet_by_name('Sheet1').cell(row=4,column=1).value)
